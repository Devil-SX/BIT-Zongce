import os
from pathlib import Path
from openpyxl import load_workbook, Workbook

class_id="05962007"

source_dir = Path('./process_5')
root_dir = Path(f'./{class_id}班材料')
summerize_excel = root_dir / '（2023-2版）2020级表2：北京理工大学信息与电子学院个人及班级综测成绩汇总版.xlsx'
# 以上目录都以成功创建好


# 删除原有文件
if os.path.exists(summerize_excel):
    os.remove(summerize_excel)
    print("[Info] Old file removed")

# 汇总表格
# 创建一个新的工作簿
new_wb = Workbook()
new_ws = new_wb.active

i = 1

# 遍历当前目录下的所有子目录
for subdir, dirs, files in os.walk(str(source_dir)):
    # 遍历子目录中的所有文件
    for filename in files:
        # 如果文件是Excel文件
        if filename.endswith('.xlsx') and (("表2" in filename) or ("表二" in filename)):
            # 加载工作簿
            wb = load_workbook(os.path.join(subdir, filename))
            # 获取第一个工作表
            ws = wb.active
            # 有的同学填在第四行，有的同学填在第五行
            # 优先级 5行 > 4行
            row_data = [cell.value for cell in ws[5]]
            if not (row_data[0] or row_data[1] or row_data[2]):
                row_data = [cell.value for cell in ws[4]]
            
            # row_data.append(subdir) # For Test
            row_data[0] = str(i) # 序号
            row_data[3] = str(row_data[3]) # 学号
            if str(row_data[4]) != class_id:
                print("[Warning] Class ID not match!")
                print(f"Wrong ID is {row_data[4]}")
                print(f"from {subdir}")
            row_data[4] = class_id # 班级
            # print(row_data)

            # 将数据添加到新工作表中
            new_ws.append(row_data)
            i = i + 1

# 保存新工作簿
new_wb.save(summerize_excel)
